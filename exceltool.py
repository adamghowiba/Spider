import openpyxl
import fileutil
import configutil
import hunter


class ExcelTool:

    def __init__(self):
        self.excel_path = fileutil.get_project_file("Spider", "info.xlsx")


        # Create workbook
        self.wb = openpyxl.load_workbook(self.excel_path)

        # Fetches the current active workbook
        self.ws = self.wb.active

        self.headers = ["Email", "First Name", "Last Name", "Phone Number", "Type"]
        self.add_headers()
        self.col = 1
        self.row = configutil.get_last_row()

    def add_headers(self):
        headers_row = 1
        for header in self.headers:
            self.ws.cell(1, headers_row, value=header)
            headers_row += 1

    def add_company_data(self, company):
        company = hunter.get_company_json(company)

        row_difference = 1
        for data in company["emails"]:
            self.ws.cell(row=self.row, column=self.col, value=data['value'])
            self.ws.cell(row=self.row, column=self.col + 1, value=data["first_name"])
            self.ws.cell(row=self.row, column=self.col + 2, value=data["last_name"])
            self.ws.cell(row=self.row, column=self.col + 3, value=data["phone_number"])
            self.ws.cell(row=self.row, column=self.col + 4, value=data["type"])
            row_difference += 1
            self.row += 1
        configutil.add_last_row(row_difference)
        print(F"Spider added {row_difference} new records")
        self.save_file()

    def save_file(self):
        self.wb.save(self.excel_path)
        # print("Saved File")

        # print("Last row ended on:", self.last_row)
